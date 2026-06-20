import io
from datetime import datetime

from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import SECRET_KEY
from models import SessionLocal, VanurRecord, init_db

app = Flask(__name__)
app.secret_key = SECRET_KEY


@app.route("/")
def index():
    return render_template("index.html", active_page="entry")


@app.route("/save", methods=["POST"])
def save():
    part_number = request.form.get("part_number", "").strip()
    blo_name = request.form.get("blo_name", "").strip()
    blo_designation = request.form.get("blo_designation", "").strip()
    blo_mobile = request.form.get("blo_mobile", "").strip()

    if not all([part_number, blo_name, blo_designation, blo_mobile]):
        flash("All fields are required.", "error")
        return redirect(url_for("index"))

    session = SessionLocal()
    try:
        record = VanurRecord(
            part_number=part_number,
            blo_name=blo_name,
            blo_designation=blo_designation,
            blo_mobile=blo_mobile,
        )
        session.add(record)
        session.commit()
        flash("Record saved successfully.", "success")
    except Exception as exc:
        session.rollback()
        flash(f"Failed to save record: {exc}", "error")
    finally:
        session.close()

    return redirect(url_for("index"))


@app.route("/api/save", methods=["POST"])
def api_save():
    data = request.get_json(silent=True) or {}
    part_number = str(data.get("part_number", "")).strip()
    blo_name = str(data.get("blo_name", "")).strip()
    blo_designation = str(data.get("blo_designation", "")).strip()
    blo_mobile = str(data.get("blo_mobile", "")).strip()

    if not all([part_number, blo_name, blo_designation, blo_mobile]):
        return jsonify({"success": False, "message": "All fields are required."}), 400

    session = SessionLocal()
    try:
        record = VanurRecord(
            part_number=part_number,
            blo_name=blo_name,
            blo_designation=blo_designation,
            blo_mobile=blo_mobile,
        )
        session.add(record)
        session.commit()
        return jsonify({"success": True, "message": "Record saved successfully.", "id": record.id})
    except Exception as exc:
        session.rollback()
        return jsonify({"success": False, "message": f"Failed to save record: {exc}"}), 500
    finally:
        session.close()


@app.route("/report")
def report():
    session = SessionLocal()
    try:
        records = session.query(VanurRecord).order_by(VanurRecord.created_at.desc()).all()
    finally:
        session.close()
    return render_template("report.html", records=records, active_page="report")


@app.route("/export")
def export_xlsx():
    session = SessionLocal()
    try:
        records = session.query(VanurRecord).order_by(VanurRecord.created_at.desc()).all()
    finally:
        session.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "AC073-VANUR Report"

    headers = ["S.No", "Part Number", "BLO Name", "BLO Designation", "BLO Mobile Number", "Created At"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=record.part_number)
        ws.cell(row=row_idx, column=3, value=record.blo_name)
        ws.cell(row=row_idx, column=4, value=record.blo_designation)
        ws.cell(row=row_idx, column=5, value=record.blo_mobile)
        ws.cell(row=row_idx, column=6, value=record.created_at.strftime("%Y-%m-%d %H:%M"))

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"AC073-VANUR_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    init_db()
    # host=0.0.0.0 allows access from phone on same Wi-Fi network
    app.run(debug=True, host="0.0.0.0", port=5000)
